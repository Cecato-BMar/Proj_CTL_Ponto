import customtkinter as ctk    
from datetime import datetime, timedelta
import openpyxl                
from tkinter import messagebox 


FORMATDATA = '%d/%m/%Y'
FORMATHORA = '%H:%M:%S'


class AppPonto:
    def __init__(self):
        self.gerar_excel("Controle de Ponto.xlsx")

        self.d_ent = []
        self.h_ent = []
        self.d_saida = []
        self.h_saida = []
        self.exib_h_ent = datetime
        self.exib_h_saida = datetime
        self.horas_trab = []

        #Criar a janela principal
        self.janela = ctk.CTk()
        self.janela.geometry("500x400")
        self.janela.title("**BMAR**")

        #Customizar janela 
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("green")

        # Design da janela
        self.texto = ctk.CTkLabel(self.janela, text= "Controle de Ponto**BMAR**", font=("Arial", 20))
        self.texto.pack(padx=10, pady=10)

        # Criar frames para entrada e saída
        self.frame_entrada = ctk.CTkFrame(master=self.janela)
        self.frame_entrada.pack(pady=20)
        self.frame_saida = ctk.CTkFrame(master=self.janela)
        self.frame_saida.pack(pady=20)
        self.frame_gravar = ctk.CTkFrame(master=self.janela)
        self.frame_gravar.pack(pady=20)

        # Criar as variáveis para as últimas datas
        self.ultima_d_ent = ctk.StringVar()
        self.ultima_d_saida = ctk.StringVar()
        self.ultima_h_ent = ctk.StringVar()
        self.ultima_h_saida = ctk.StringVar()
        self.ht = ctk.StringVar()
        self.saldo_h = ctk.StringVar()

        # Botão de entrada e label
        self.btn_entrada = ctk.CTkButton(master=self.frame_entrada, text="Entrada", command=self.entrada)
        
        self.btn_entrada.pack(side="left", padx=5)
        self.label_ent = ctk.CTkLabel(master=self.frame_entrada, text="***ENTRADA***", fg_color="green", corner_radius=20)
        
        self.label_ent.pack(side="left", padx=15)

        # Botão de saída e label
        self.btn_saida = ctk.CTkButton(master=self.frame_saida, text="Saída", command=self.saida)
        self.btn_saida.pack(side="left", padx=5)
        self.label_saida = ctk.CTkLabel(master=self.frame_saida, text="****SAÍDA****", fg_color="red", corner_radius=20)
        self.label_saida.pack(side="left", padx=15)

        #Criar um botão para salvar
        self.btn_gravar = ctk.CTkButton(master=self.frame_gravar, text="Gravar", command=self.gravar)
        self.btn_gravar.pack(pady=30 )        

        # Label horas trabalhadas no dia
        self.label_htrabalhadas = ctk.CTkLabel(master=self.frame_gravar, text="****Horas Trabalhadas****", fg_color="red", corner_radius=20)
        self.label_htrabalhadas.pack(side="left", padx=15, pady=10)

        # Label horas total
        self.label_totais = ctk.CTkLabel(master=self.frame_gravar, text="******Horas Totais******", fg_color="teal", corner_radius=20)
        self.label_totais.pack(side="right", padx=15, pady=10)
        self.janela.mainloop()

    def gerar_excel(self, caminho:str):
        '''
        Carrega um arquivo excel.
        '''
        self.workbook = openpyxl.Workbook()
        self.workbook = openpyxl.load_workbook(caminho)

        # Selecionando uma página
        self.pag = self.workbook.active

        # Criando as colunas
        self.pag['A1'] = 'Dia Entrada'
        self.pag['B1'] = 'Hora entrada'
        self.pag['C1'] = 'Dia saíada'
        self.pag['D1'] = 'Hora Saída'
        self.pag['E1'] = 'Horas Trabalhadas' 

    def pegar_datahora_atual(self) -> list[datetime, datetime]:
        '''
        Retorna a data e hora atual.
        '''
        dt = datetime.now()
        return [dt.strftime(FORMATHORA), dt.strftime(FORMATDATA)]

    def entrada(self) -> None:    
        '''
        Retorna a entrada.
        '''
        exib_h_ent, exib_d_ent = self.pegar_datahora_atual()

        print(f"Dia entrada: {exib_d_ent} --- Hora entrada: {exib_h_ent}")

        self.ultima_d_ent.set(exib_d_ent) # atualizar a variável com a última data
        self.label_ent.configure(textvariable=self.ultima_d_ent) # atualizar o rótulo com a última data

        self.ultima_h_ent.set(exib_h_ent) # atualizar a variável com a última hora
        self.label_ent.configure(textvariable=self.ultima_h_ent) # atualizar o rótulo com a última hora

        self.d_ent.append(exib_d_ent)    
        self.h_ent.append(exib_h_ent)
        
    def saida(self) -> None:  
        '''
        ??
        '''  
        exib_h_saida, exib_d_saida = self.pegar_datahora_atual()

        print(f"Dia saída: {exib_d_saida} --- Hora saída: {exib_h_saida}")
        self.ultima_d_saida.set(exib_d_saida) # atualizar a variável com a última data
        self.label_saida.configure(textvariable=self.ultima_d_saida) # atualizar o rótulo com a última data
        self.ultima_h_saida.set(exib_h_saida) # atualizar a variável com a última data
        self.label_saida.configure(textvariable=self.ultima_h_saida) # atualizar o rótulo com a última data 
        self.d_saida.append(exib_d_saida)
        self.h_saida.append(exib_h_saida)  
        
    def horas_trabalhadas(self) -> None:    
        '''
        Altera a janela com a quantidade total de horas trabalhadas. 
        '''
        for i in range(len(self.h_ent)):       
            horas = datetime.strptime(self.h_saida[i], FORMATHORA) - datetime.strptime(self.h_ent[i], FORMATHORA)
            self.horas_trab.append(horas)   
            print(self.horas_trab)

        total_horas = sum(self.horas_trab, timedelta())    
        print(total_horas)
        self.ht.set(total_horas) # atualizar a variável com a última data
        self.label_htrabalhadas.configure(textvariable=self.ht) # atualizar o rótulo com a última data
        self.ht.set(total_horas) # atualizar a variável com a última data 
        
    def gravar(self) -> None: 
        '''
        Grava o horario de saída em um arquivo excel.
        '''
        self.horas_trabalhadas()   

        for i in range(len(self.d_ent)):         
            self.pag.append((self.d_ent[i], self.h_ent[i], self.d_saida[i], self.h_saida[i], self.horas_trab[i]))
            
        self.workbook.save('Controle de Ponto.xlsx')             
        messagebox.showinfo("Dados Salvos", "Os dados foram salvos com sucesso!")            
        print(self.total_horas())
        
    def total_horas(self):
        '''
        ??
        '''
        total_timedelta = timedelta()

        for row in self.pag.iter_rows(min_row=2, values_only=True):
            valor_celula = row[5 - 1]  # Subtrai 1 porque as colunas no Excel são indexadas a partir de 1
            if isinstance(valor_celula, timedelta):
                total_timedelta += valor_celula
        self.saldo_h.set(total_timedelta) # atualizar a variável com a última data
        self.label_totais.configure(textvariable=self.saldo_h) # atualizar o rótulo com a última data
        self.saldo_h.set(total_timedelta) # atualizar a variável com a última data 
        return total_timedelta 


if __name__ == '__main__':
    AppPonto()