# Projeto edstinado a controlar o ponto de um funcionário no trabalho.
from datetime import datetime
import openpyxl
from tkinter import *
from tkinter import ttk


#criando um arquivo excell 
workbook = openpyxl.Workbook()
workbook = openpyxl.load_workbook('Controle de Ponto.xlsx')
# Selecionando uma página
pag = workbook.active
# Criando as colunas
pag['A1'] = 'Entrada'
pag['B1'] = 'Saída'

 #1- Pegar a data e hora
ent = []
saida = []
dt = datetime.now()
dt = (dt.strftime('%d /%m/ %Y %H:%M:%S '))
print(dt)


def painel():
    print("=" * 30)
    print("PROGRAMA CONTROLE DE PONTO")
    print("=" * 30)
    print("1- Entrada.")
    print("2- Saída.")
    print("3- Sair do programa.")
    opcao = int(input("Opção : "))
    ponto(opcao)
    print("=" * 30)
    print("-"*13 + "FIM" + "-"*13 )
    print("=" * 30)
    

#2- Atribuir a entrada ou saída
   
def ponto(opcao):
     while opcao != 3 :
        if opcao < 1 or opcao > 3:
            print("Opção inválida, tente novamente!")
            opcao = int(input("Opção : "))    
        if opcao == 1 :
            dt1 = datetime.now()
            ent.append(dt1.strftime('%d /%m/ %Y %H:%M:%S ')) 
            
            opcao = int(input("Opção : "))
            
     
        elif opcao == 2 :
            dt2 = datetime.now()
            saida.append(dt2.strftime('%d /%m/ %Y %H:%M:%S '))
            
            opcao = int(input("Opção : "))
           
            
       
     
     for i in range(len(ent)):
        dados = (ent[i],saida[i])
        pag.append(dados)
       
     exibir()
     #3- Salvar em um arquivo excell
     workbook.save('Controle de Ponto.xlsx')
     
def exibir():
    # Obtenha o número de linhas e colunas da planilha
    num_linhas = pag.max_row
    num_colunas = pag.max_column
    print("-"*5 + "ENTRADA" + "-"*5 + " "* 10 + "-"*5 + "SAÌDA" + "-"*5)
    # Itere sobre as células da planilha e imprima o valor de cada célula
    for i in range(2, num_linhas + 1):
        for j in range(1, num_colunas + 1):
            cell_value = pag.cell(row=i, column=j).value
            print(cell_value, end="\t")
        print()  # Adicione uma nova linha após cada linha de células
    
painel()


   



    





 
    




