# import tkinter

# janela = tkinter.Tk()  #instanciar u8ma janela
# janela.geometry("500x300") # define o tambanho da janela

# texto = tkinter.Label(janela, text= "Controle de Ponto**BMAR**") # cria uma variável para armazenar o texto que será exibido na label
# texto.pack(janela, padx=10, pady=10)

# btn_entrada = tkinter.Button(janela, text= "Entrada")
# btn_entrada.pack(padx=10, pady=10)





# janela.mainloop() #Faz com que a janela seja exibida.

import customtkinter
from datetime import datetime
import openpyxl


#criando um arquivo excell 
workbook = openpyxl.Workbook()
workbook = openpyxl.load_workbook('Controle de Ponto.xlsx')
# Selecionando uma página
pag = workbook.active
# Criando as colunas
pag['A1'] = 'Entrada'
pag['B1'] = 'Saída'
 #1- Pegar a data e hora


ent = []
saida = []
dt = datetime.now()
dt = (dt.strftime('%d /%m/ %Y %H:%M:%S '))

customtkinter.set_appearance_mode("dark")  
customtkinter.set_default_color_theme("green")


janela = customtkinter.CTk()
janela.geometry("500x300")



def entrada():
    dt1 = datetime.now()
    ent.append(dt1.strftime('%d /%m/ %Y %H:%M:%S ')) 
    print(dt1)
    # Obter a próxima linha vazia na planilha
    #prox_linha = pag.max_row + 1

    # Inserir as informações de entrada e saída na planilha
   # pag.cell(row=prox_linha, column=1, value=ent[-1])
    #pag.cell(row=prox_linha, column=2, value=saida[-1])

    # Salvar as alterações no arquivo
    workbook.save('Controle de Ponto.xlsx')

def bsaida():
    dt2 = datetime.now()
    saida.append(dt2.strftime('%d /%m/ %Y %H:%M:%S '))
    print(dt2)  
    # Obter a próxima linha vazia na planilha
    #prox_linha = pag.max_row + 1

    # Inserir as informações de entrada e saída na planilha
    #pag.cell(row=prox_linha, column=1, value=ent[-1])
    #pag.cell(row=prox_linha, column=2, value=saida[-1])

    # Salvar as alterações no arquivo
    workbook.save('Controle de Ponto.xlsx')
    
def gravar():
    texto1 = entrada 
    texto2 = customtkinter.CTkLabel(janela, text=  texto1)
    texto2.pack(padx=10, pady=10)
    
#     num_linhas = pag.max_row
#     num_colunas = pag.max_column

#     for i in range(2, num_linhas + 1):
#         for j in range(1, num_colunas + 1):
#             cell_value = pag.cell(row=i, column=j).value
#             print(cell_value, end="\t")
#             janela.combobox = customtkinter.CTkComboBox(master=janela, values=[cell_value])
#             janela.combobox.grid(row=1, column=0, padx=20, pady=20)
#             janela.textbox.insert(cell_value)

#     workbook.save('Controle de Ponto.xlsx')
             
texto = customtkinter.CTkLabel(janela, text= "Controle de Ponto**BMAR**")
texto.pack(padx=10, pady=10)

btn_entrada = customtkinter.CTkButton(janela, text= "Entrada" , command= entrada)
btn_entrada.pack(padx=10, pady=10)

btn_saida = customtkinter.CTkButton(janela, text= "Saída" , command= bsaida)
btn_saida.pack(padx=10, pady=10)

btn_gravar = customtkinter.CTkButton(janela, text= "Gravar" , command= gravar)
btn_gravar.pack(padx=10, pady=10)

#Criar o widget de lista
listbox = customtkinter.Listbox(janela, width=30)

#Criar a lista com os dados da tabela

#Inserir cada linha da tabela como um novo item na lista
for i in range(len(ent)):
     dados = (ent[i],saida[i])
     pag.append(dados)
     listbox.insert(janela.END, "{:<10} {:<20} {:>10}".format(*dados))

# Adicionar a lista à janela e exibir
listbox.pack()

table = customtkinter.Table(janela, headings=["Entrada", "Saída"], rows=[(ent, saida)])
     
table.pack()
#Criar a tabela
tabela = customtkinter.CTkTabview(janela, headings=["Entrada", "Saída"], rows=zip(ent, saida))

#Adicionar a tabela à janela
tabela.pack(padx=10, pady=10)

janela.mainloop()